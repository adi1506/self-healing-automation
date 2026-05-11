from __future__ import annotations
import os
import re
import yaml
from datetime import datetime
from openpyxl import load_workbook
from core.scenarios import Scenario, save_scenario, ScenarioValidationError

MARKER_NAME = "_migrated.yaml"


def _read_marker(scenarios_dir: str) -> dict:
    p = os.path.join(scenarios_dir, MARKER_NAME)
    if not os.path.exists(p):
        return {"recipes": [], "flows": [], "test_data": []}
    with open(p, encoding="utf-8") as f:
        return yaml.safe_load(f) or {"recipes": [], "flows": [], "test_data": []}


def _write_marker(scenarios_dir: str, marker: dict) -> None:
    os.makedirs(scenarios_dir, exist_ok=True)
    with open(os.path.join(scenarios_dir, MARKER_NAME), "w", encoding="utf-8") as f:
        yaml.safe_dump(marker, f, sort_keys=False)


def _slug(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_-]+", "_", name).strip("_")
    return s or "scenario"


def _migrate_recipe_file(path: str, scenarios_dir: str) -> bool:
    with open(path, encoding="utf-8") as f:
        recipe = yaml.safe_load(f) or {}
    if not recipe.get("name") or not recipe.get("steps"):
        return False
    sc = Scenario(
        id=_slug(recipe["name"]),
        name=recipe["name"],
        kind="single-page",
        base_url=recipe.get("start_url", ""),
        steps=recipe["steps"],
        dataset=[],
        expected_outcome=recipe.get("expected_outcome", "success"),
        assertions=recipe.get("assertions", []),
        created_at=datetime.now().isoformat(timespec="seconds"),
    )
    try:
        save_scenario(scenarios_dir, sc)
    except ScenarioValidationError:
        return False
    return True


def _migrate_flow_file(path: str, scenarios_dir: str) -> bool:
    with open(path, encoding="utf-8") as f:
        flow = yaml.safe_load(f) or {}
    if not flow.get("name") or not flow.get("recipes"):
        return False
    sc = Scenario(
        id=_slug(flow["name"]),
        name=flow["name"],
        kind="multi-page",
        base_url="",
        steps=[],
        dataset=[],
        expected_outcome=flow.get("expected_outcome", "success"),
        recipe_refs=list(flow["recipes"]),
        created_at=datetime.now().isoformat(timespec="seconds"),
    )
    try:
        save_scenario(scenarios_dir, sc)
    except ScenarioValidationError:
        return False
    return True


def _migrate_test_data(xlsx_path: str, scenarios_dir: str) -> bool:
    wb = load_workbook(xlsx_path)
    if "Test Data" not in wb.sheetnames or "Element Map" not in wb.sheetnames:
        return False
    em = wb["Element Map"]
    td = wb["Test Data"]
    em_rows = list(em.iter_rows(values_only=True))
    td_rows = list(td.iter_rows(values_only=True))
    if len(td_rows) <= 1:
        return False  # no data rows

    # Build editable field list from Element Map (skip buttons)
    em_header = em_rows[0]
    name_col = em_header.index("Element Name") if "Element Name" in em_header else 1
    type_col = em_header.index("Element Type") if "Element Type" in em_header else 2
    editable_fields = [r[name_col] for r in em_rows[1:]
                       if r[type_col] != "button" and r[name_col]]

    td_header = td_rows[0]
    field_to_col = {h: i for i, h in enumerate(td_header) if h}

    dataset: list[dict] = []
    for row in td_rows[1:]:
        if all(cell in (None, "") for cell in row):
            continue
        rec = {}
        for fname in editable_fields:
            if fname in field_to_col:
                v = row[field_to_col[fname]]
                rec[fname] = "" if v is None else v
        if "Expected Outcome" in field_to_col:
            v = row[field_to_col["Expected Outcome"]]
            rec["__expected_outcome"] = v or "success"
        dataset.append(rec)

    if not dataset:
        return False

    base_slug = _slug(os.path.splitext(os.path.basename(xlsx_path))[0])
    fill_steps = [{"action": "fill", "target": f, "value": f"{{{{{f}}}}}"}
                  for f in editable_fields]
    # Use a placeholder URL so single-page validation passes; user updates via Settings.
    placeholder_url = f"migrated://{base_slug}"
    sc = Scenario(
        id=base_slug + "_data",
        name=f"{base_slug} — data-driven",
        kind="single-page",
        base_url=placeholder_url,
        steps=fill_steps or [{"action": "wait_for_url", "contains": ""}],
        dataset=dataset,
        expected_outcome="success",
        created_at=datetime.now().isoformat(timespec="seconds"),
    )
    try:
        save_scenario(scenarios_dir, sc)
    except ScenarioValidationError:
        return False
    return True


def migrate_all(
    recipes_dir: str, flows_dir: str, scans_dir: str, scenarios_dir: str,
) -> dict:
    marker = _read_marker(scenarios_dir)
    report = {"recipes_migrated": 0, "flows_migrated": 0, "test_data_migrated": 0}

    if os.path.isdir(recipes_dir):
        for f in sorted(os.listdir(recipes_dir)):
            if not f.endswith(".yaml") or f in marker["recipes"]:
                continue
            if _migrate_recipe_file(os.path.join(recipes_dir, f), scenarios_dir):
                marker["recipes"].append(f)
                report["recipes_migrated"] += 1

    if os.path.isdir(flows_dir):
        for f in sorted(os.listdir(flows_dir)):
            if not f.endswith(".yaml") or f in marker["flows"]:
                continue
            if _migrate_flow_file(os.path.join(flows_dir, f), scenarios_dir):
                marker["flows"].append(f)
                report["flows_migrated"] += 1

    if os.path.isdir(scans_dir):
        for f in sorted(os.listdir(scans_dir)):
            if not f.endswith(".xlsx") or f in marker["test_data"]:
                continue
            if _migrate_test_data(os.path.join(scans_dir, f), scenarios_dir):
                marker["test_data"].append(f)
                report["test_data_migrated"] += 1

    _write_marker(scenarios_dir, marker)
    return report
