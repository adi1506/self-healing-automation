import os
import re
from datetime import datetime
from zipfile import BadZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


# Status color mapping
STATUS_FILLS = {
    "NEW": PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid"),       # Green
    "CHANGED": PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),     # Yellow
    "REMOVED": PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid"),    # Red
    "UNRESOLVED": PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid"), # Orange
}

ELEMENT_MAP_HEADERS = [
    "S.No", "Element Name", "Element Type",
    "Locator ID", "Locator Name", "Locator CSS", "Locator XPath",
    "Locator Data-TestID", "Locator Label",
    "Placeholder", "Available Options", "Current Value",
    "Status", "Change Details", "Healed By", "Last Scanned",
    "Pattern", "Title", "Min Length", "Max Length",
    "Min Value", "Max Value", "Required", "Autocomplete", "Helper Text",
]

LAST_SCANNED_COL = 16

# Position-aligned keys for Element Map; None marks the Last Scanned column
# which is written separately as a timestamp.
EXTENDED_ELEMENT_MAP_KEYS = [
    "sno", "element_name", "element_type",
    "locator_id", "locator_name", "locator_css", "locator_xpath",
    "locator_data_testid", "locator_label",
    "placeholder", "available_options", "current_value",
    "status", "change_details", "healed_by",
    None,  # column 16 = "Last Scanned"
    "pattern", "title_attr", "minlength", "maxlength",
    "min", "max", "required", "autocomplete", "helper_text",
]

# Kept for any callers that iterate the original 15 keys
ELEMENT_MAP_KEYS = [k for k in EXTENDED_ELEMENT_MAP_KEYS[:15]]

RUN_RESULTS_HEADERS = [
    "Run ID", "Timestamp", "Test Case Name", "Element Name",
    "Expected Value", "Actual Value", "Status", "Screenshot",
]

SCAN_HISTORY_HEADERS = [
    "Scan ID", "Timestamp", "Total Elements", "New", "Changed", "Removed", "Unchanged",
]

HEAL_HISTORY_HEADERS = [
    "Heal ID", "Timestamp", "Element Name", "Change Type", "Change Details", "Healed By",
]

PAGE_CONTEXT_HEADERS = ["Title", "H1", "First Paragraph"]

# Element types that are NOT editable (excluded from Test Data sheet)
NON_EDITABLE_TYPES = {"button"}


class ExcelManager:
    def __init__(self, data_dir="data/scans"):
        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)
        self._url_map_file = os.path.join(data_dir, "_url_map.txt")

    def sanitize_url(self, url: str) -> str:
        """Convert URL to a safe filename string. Strips fragment (#hash) so all versions share one file."""
        url_no_fragment = re.sub(r"#.*$", "", url)
        sanitized = re.sub(r"https?://", "", url_no_fragment)
        sanitized = re.sub(r"[^a-zA-Z0-9]", "_", sanitized)
        sanitized = re.sub(r"_+", "_", sanitized).strip("_")
        return sanitized

    def get_excel_path(self, url: str) -> str:
        """Get the Excel file path for a given URL."""
        return os.path.join(self.data_dir, f"{self.sanitize_url(url)}.xlsx")

    def _save_workbook(self, wb, path: str):
        """Atomically save a workbook: write to a temp file, then replace.
        Prevents readers from seeing a half-written zip if a save is interrupted
        or read concurrently."""
        tmp_path = path + ".tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, path)

    def _load_workbook(self, path: str):
        """Load a workbook with a clearer error if the file is corrupt."""
        try:
            return load_workbook(path)
        except BadZipFile as e:
            raise RuntimeError(
                f"Excel file is corrupt or empty: {path}. "
                "Re-scan the URL from the Scanner page to regenerate it."
            ) from e

    def excel_exists(self, url: str) -> bool:
        """Check if an Excel file exists for the given URL."""
        return os.path.exists(self.get_excel_path(url))

    def _strip_fragment(self, url: str) -> str:
        """Strip the URL fragment (#hash) so all versions share one entry."""
        return re.sub(r"#.*$", "", url)

    def _save_url_mapping(self, url: str):
        """Save URL to sanitized-name mapping for reverse lookup."""
        base_url = self._strip_fragment(url)
        mappings = {}
        if os.path.exists(self._url_map_file):
            with open(self._url_map_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if "|" in line:
                        key, val = line.split("|", 1)
                        mappings[key] = val
        mappings[self.sanitize_url(base_url)] = base_url
        with open(self._url_map_file, "w") as f:
            for key, val in mappings.items():
                f.write(f"{key}|{val}\n")

    def delete_url(self, url: str) -> bool:
        """Delete all data for a scanned URL (Excel file + URL mapping entry)."""
        base_url = self._strip_fragment(url)
        sanitized = self.sanitize_url(base_url)

        # Delete the Excel file
        excel_path = self.get_excel_path(url)
        if os.path.exists(excel_path):
            os.remove(excel_path)

        # Remove from URL map
        if os.path.exists(self._url_map_file):
            mappings = {}
            with open(self._url_map_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if "|" in line:
                        key, val = line.split("|", 1)
                        mappings[key] = val
            if sanitized in mappings:
                del mappings[sanitized]
                with open(self._url_map_file, "w") as f:
                    for key, val in mappings.items():
                        f.write(f"{key}|{val}\n")
                return True
        return not os.path.exists(excel_path)

    def list_scanned_urls(self) -> list[str]:
        """List all URLs that have been scanned (have Excel files)."""
        if not os.path.exists(self._url_map_file):
            return []
        urls = []
        with open(self._url_map_file, "r") as f:
            for line in f:
                line = line.strip()
                if "|" in line:
                    _, url = line.split("|", 1)
                    if os.path.exists(self.get_excel_path(url)):
                        urls.append(url)
        return urls

    def save_element_map(self, url: str, elements: list[dict]) -> str:
        """Save scanned elements to Excel. Creates or overwrites the Element Map sheet."""
        path = self.get_excel_path(url)
        self._save_url_mapping(url)

        wb = None
        if os.path.exists(path):
            try:
                wb = load_workbook(path)
            except BadZipFile:
                wb = None
        if wb is None:
            wb = Workbook()
            wb.remove(wb.active)

        if "Element Map" in wb.sheetnames:
            del wb["Element Map"]
        ws = wb.create_sheet("Element Map", 0)

        for col, header in enumerate(ELEMENT_MAP_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row_idx, elem in enumerate(elements, 2):
            for col_idx, key in enumerate(EXTENDED_ELEMENT_MAP_KEYS, 1):
                if key is None:
                    continue
                ws.cell(row=row_idx, column=col_idx, value=elem.get(key, ""))
            ws.cell(row=row_idx, column=LAST_SCANNED_COL, value=timestamp)

            status = elem.get("status", "")
            if status in STATUS_FILLS:
                fill = STATUS_FILLS[status]
                for col_idx in range(1, len(ELEMENT_MAP_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        if "Test Data" not in wb.sheetnames:
            ws_td = wb.create_sheet("Test Data")
            ws_td.cell(row=1, column=1, value="S.No")
            ws_td.cell(row=1, column=2, value="Test Case Name")
            ws_td.cell(row=1, column=3, value="AI Context")
            col = 4
            for elem in elements:
                if elem.get("element_type") not in NON_EDITABLE_TYPES:
                    ws_td.cell(row=1, column=col, value=elem["element_name"])
                    col += 1
        else:
            ws_td = wb["Test Data"]
            # Backfill AI Context column for older sheets that don't have it
            if ws_td.cell(row=1, column=3).value != "AI Context":
                ws_td.insert_cols(3)
                ws_td.cell(row=1, column=3, value="AI Context")
            existing_headers = []
            for col in range(3, ws_td.max_column + 1):
                val = ws_td.cell(row=1, column=col).value
                if val:
                    existing_headers.append(val)
            new_editable_names = [
                e["element_name"] for e in elements
                if e.get("element_type") not in NON_EDITABLE_TYPES
            ]
            next_col = ws_td.max_column + 1
            for name in new_editable_names:
                if name not in existing_headers:
                    ws_td.cell(row=1, column=next_col, value=name)
                    next_col += 1

        for sheet_name, headers in [
            ("Run Results", RUN_RESULTS_HEADERS),
            ("Scan History", SCAN_HISTORY_HEADERS),
            ("Heal History", HEAL_HISTORY_HEADERS),
        ]:
            if sheet_name not in wb.sheetnames:
                ws_other = wb.create_sheet(sheet_name)
                for col, header in enumerate(headers, 1):
                    ws_other.cell(row=1, column=col, value=header)

        self._save_workbook(wb, path)
        return path

    def read_element_map(self, url: str) -> list[dict]:
        """Read the Element Map sheet and return as list of dicts."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return []

        wb = self._load_workbook(path)
        ws = wb["Element Map"]
        elements = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                break
            elem = {}
            for col_idx, key in enumerate(EXTENDED_ELEMENT_MAP_KEYS, 1):
                if key is None:
                    continue
                raw = ws.cell(row=row, column=col_idx).value
                elem[key] = raw if raw is not None else ""
            # Last Scanned at column 16
            raw_ts = ws.cell(row=row, column=LAST_SCANNED_COL).value
            elem["last_scanned"] = raw_ts if raw_ts is not None else ""
            # Coerce booleans for `required` (it round-trips through Excel as True/False or "")
            req = elem.get("required", "")
            elem["required"] = bool(req) if req != "" else False
            elements.append(elem)
        return elements

    def save_test_data(self, url: str, test_rows: list[dict]):
        """Save test data rows to the Test Data sheet."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return
        wb = self._load_workbook(path)
        ws = wb["Test Data"]

        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)

        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col, value=None)

        for row_idx, test_row in enumerate(test_rows, 2):
            for col_idx, header in enumerate(headers, 1):
                key = header.lower().replace(" ", "_").replace(".", "")
                value = test_row.get(key) or test_row.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        self._save_workbook(wb, path)

    def read_test_data(self, url: str) -> list[dict]:
        """Read test data rows from the Test Data sheet."""
        path = self.get_excel_path(url)
        wb = self._load_workbook(path)
        ws = wb["Test Data"]

        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)

        rows = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                break
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                raw = ws.cell(row=row, column=col_idx).value
                row_data[header] = raw if raw is not None else ""
            rows.append(row_data)
        return rows

    def _normalize_key(self, s: str) -> str:
        """Normalize a header or data key for comparison."""
        return s.lower().replace(" ", "_").replace("-", "").replace(".", "")

    def _append_to_sheet(self, url: str, sheet_name: str, data: dict, headers: list[str]):
        """Generic method to append a row to any sheet."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return
        wb = self._load_workbook(path)
        ws = wb[sheet_name]

        next_row = 2
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=1).value is None:
                next_row = row
                break

        keys = [self._normalize_key(h) for h in headers]
        for col_idx, key in enumerate(keys, 1):
            for data_key, data_val in data.items():
                if self._normalize_key(data_key) == key:
                    ws.cell(row=next_row, column=col_idx, value=data_val)
                    break

        self._save_workbook(wb, path)

    def append_run_result(self, url: str, result: dict):
        """Append a run result row to the Run Results sheet."""
        self._append_to_sheet(url, "Run Results", result, RUN_RESULTS_HEADERS)

    def read_run_results(self, url: str) -> list[dict]:
        """Read all run results."""
        return self._read_sheet(url, "Run Results", RUN_RESULTS_HEADERS)

    def append_scan_history(self, url: str, entry: dict):
        """Append a scan history entry."""
        self._append_to_sheet(url, "Scan History", entry, SCAN_HISTORY_HEADERS)

    def read_scan_history(self, url: str) -> list[dict]:
        """Read all scan history entries."""
        return self._read_sheet(url, "Scan History", SCAN_HISTORY_HEADERS)

    def append_heal_history(self, url: str, entry: dict):
        """Append a heal history entry."""
        self._append_to_sheet(url, "Heal History", entry, HEAL_HISTORY_HEADERS)

    def read_heal_history(self, url: str) -> list[dict]:
        """Read all heal history entries."""
        return self._read_sheet(url, "Heal History", HEAL_HISTORY_HEADERS)

    def save_page_context(self, url: str, ctx: dict) -> None:
        """Save page-level context (title, h1, first paragraph) to the Page Context sheet."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return
        wb = self._load_workbook(path)
        if "Page Context" in wb.sheetnames:
            del wb["Page Context"]
        ws = wb.create_sheet("Page Context")
        for col, header in enumerate(PAGE_CONTEXT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        ws.cell(row=2, column=1, value=ctx.get("title", ""))
        ws.cell(row=2, column=2, value=ctx.get("h1", ""))
        ws.cell(row=2, column=3, value=ctx.get("first_paragraph", ""))
        self._save_workbook(wb, path)

    def read_page_context(self, url: str) -> dict:
        """Read page-level context from the Page Context sheet."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return {"title": "", "h1": "", "first_paragraph": ""}
        wb = self._load_workbook(path)
        if "Page Context" not in wb.sheetnames:
            return {"title": "", "h1": "", "first_paragraph": ""}
        ws = wb["Page Context"]
        return {
            "title": ws.cell(row=2, column=1).value or "",
            "h1": ws.cell(row=2, column=2).value or "",
            "first_paragraph": ws.cell(row=2, column=3).value or "",
        }

    def _read_sheet(self, url: str, sheet_name: str, headers: list[str]) -> list[dict]:
        """Generic method to read all rows from a sheet."""
        path = self.get_excel_path(url)
        if not os.path.exists(path):
            return []

        wb = self._load_workbook(path)
        ws = wb[sheet_name]
        keys = [self._normalize_key(h) for h in headers]

        rows = []
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value is None:
                break
            row_data = {}
            for col_idx, key in enumerate(keys, 1):
                raw = ws.cell(row=row, column=col_idx).value
                row_data[key] = raw if raw is not None else ""
            rows.append(row_data)
        return rows
