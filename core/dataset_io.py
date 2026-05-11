from __future__ import annotations
import io
import csv
from openpyxl import Workbook, load_workbook


def _drop_blank(rows: list[dict]) -> list[dict]:
    return [r for r in rows if any((v or "").strip() for v in r.values())]


def parse_csv_bytes(blob: bytes) -> list[dict]:
    text = blob.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return _drop_blank([
        {k: ("" if v is None else str(v)) for k, v in r.items() if k}
        for r in reader
    ])


def parse_xlsx_bytes(blob: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(blob), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        rec = {h: ("" if v is None else str(v)) for h, v in zip(header, row) if h}
        out.append(rec)
    return _drop_blank(out)


def dataset_to_xlsx_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    if not rows:
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
    header = list(rows[0].keys())
    ws.append(header)
    for r in rows:
        ws.append([r.get(h, "") for h in header])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
