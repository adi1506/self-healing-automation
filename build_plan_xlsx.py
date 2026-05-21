"""Build the 6-sprint project plan Excel for manager review."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"e:\AIWork\self-healing-automation\Self_Healing_Automation_Plan.xlsx"

# (Sprint, Phase, DateRange, SubTaskHeading, Description, DemoDate, Owner, Status)
ROWS = [
    # ---------- Sprint 1 ----------
    (1, "Planning & Foundations", "Mon Apr 27 to Fri May 1",
     "Problem framing", "Agreed what self-healing means for our tool.",
     "", "Aditya", "Done"),
    (1, "Planning & Foundations", "Mon Apr 27 to Fri May 1",
     "Tech choices", "Picked Playwright, Streamlit, and a local AI model so nothing leaks externally.",
     "", "Aditya", "Done"),
    (1, "Planning & Foundations", "Mon Apr 27 to Fri May 1",
     "Storage approach", "Settled on plain text files on disk for every artifact.",
     "", "Aditya", "Done"),
    (1, "Planning & Foundations", "Mon Apr 27 to Fri May 1",
     "Anti-bot handling", "Figured out how to avoid being blocked by target sites.",
     "", "Aditya", "Done"),

    # ---------- Sprint 2 ----------
    (2, "Scanner & Crawler", "Mon May 4 to Fri May 8",
     "Page scanner", "Lists every input, button, and dropdown on a page.",
     "", "Aditya", "Done"),
    (2, "Scanner & Crawler", "Mon May 4 to Fri May 8",
     "Site crawler", "Walks a whole site and scans each page it finds.",
     "", "Aditya", "Done"),
    (2, "Scanner & Crawler", "Mon May 4 to Fri May 8",
     "Excel export", "Saves each scan as an Excel sheet for review.",
     "", "Aditya", "Done"),
    (2, "Scanner & Crawler", "Mon May 4 to Fri May 8",
     "Library screen", "Lists past scans and lets users open them.",
     "", "Aditya", "Done"),

    # ---------- Sprint 3 ----------
    (3, "Runner & Self-Healing", "Mon May 11 to Fri May 15",
     "Action runner", "Runs the steps of a scenario, clicking and typing as needed.",
     "", "Aditya", "Done"),
    (3, "Runner & Self-Healing", "Mon May 11 to Fri May 15",
     "Self-healing", "Tries alternative locators, then the AI, if the original fails.",
     "", "Aditya", "Done"),
    (3, "Runner & Self-Healing", "Mon May 11 to Fri May 15",
     "Multi-page flows", "A scenario can move across several pages.",
     "", "Aditya", "Done"),
    (3, "Runner & Self-Healing", "Mon May 11 to Fri May 15",
     "Reports view", "Shows pass and fail history and flags self-healed steps.",
     "", "Aditya", "Done"),
    (3, "Runner & Self-Healing", "Mon May 11 to Fri May 15",
     "Smart test data", "Fills fields with realistic values instead of dummy text.",
     "", "Aditya", "Done"),

    # ---------- Sprint 4 ----------
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "In-page capture", "Records the user's clicks, typing, and dropdown changes.",
     "Fri May 22", "Aditya", "In Progress"),
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "Multiple locators", "Stores several ways to find each field for safer replay.",
     "Fri May 22", "Aditya", "In Progress"),
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "Encrypted login", "Saves the login session so users do not re-login on replay.",
     "Fri May 22", "Aditya", "In Progress"),
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "Recordings screen", "Lists past recordings and plays them back.",
     "Fri May 22", "Aditya", "In Progress"),
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "Round-trip test", "A saved recording still replays after closing and reopening the tool.",
     "Fri May 22", "Aditya", "In Progress"),
    (4, "Scenario Recording", "Mon May 18 to Fri May 22",
     "MILESTONE DEMO",
     "Record a fresh scenario live, save, reload, and replay on a slightly changed page.",
     "Fri May 22", "Aditya", "Planned"),

    # ---------- Sprint 5 ----------
    (5, "AI Test Cases, Part 1", "Mon May 25 to Fri May 29",
     "Happy variants", "From one recording, generate several runs with different realistic data.",
     "Fri May 29", "Aditya", "Planned"),
    (5, "AI Test Cases, Part 1", "Mon May 25 to Fri May 29",
     "Persist cases", "Save generated cases next to the recording so they survive restarts.",
     "Fri May 29", "Aditya", "Planned"),
    (5, "AI Test Cases, Part 1", "Mon May 25 to Fri May 29",
     "Reuse runner", "Run each generated case through the existing replay engine.",
     "Fri May 29", "Aditya", "Planned"),
    (5, "AI Test Cases, Part 1", "Mon May 25 to Fri May 29",
     "MILESTONE DEMO",
     "Generate five or more variants from one recording and run them.",
     "Fri May 29", "Aditya", "Planned"),

    # ---------- Sprint 6 ----------
    (6, "AI Test Cases, Part 2", "Mon Jun 1 to Fri Jun 5",
     "Negative cases", "Generate broken inputs on purpose to test how the app handles bad data.",
     "Fri Jun 5", "Aditya", "Planned"),
    (6, "AI Test Cases, Part 2", "Mon Jun 1 to Fri Jun 5",
     "Review screen", "Tab to read, edit, approve, or discard each generated test.",
     "Fri Jun 5", "Aditya", "Planned"),
    (6, "AI Test Cases, Part 2", "Mon Jun 1 to Fri Jun 5",
     "Failure reasons", "When a test fails, say in plain English if the data was wrong or the page changed.",
     "Fri Jun 5", "Aditya", "Planned"),
    (6, "AI Test Cases, Part 2", "Mon Jun 1 to Fri Jun 5",
     "Per-case reports", "Reports screen shows pass and fail per generated test plus any self-healing.",
     "Fri Jun 5", "Aditya", "Planned"),
    (6, "AI Test Cases, Part 2", "Mon Jun 1 to Fri Jun 5",
     "MILESTONE DEMO",
     "Record once, generate cases, review and approve, run them, and show the report.",
     "Fri Jun 5", "Aditya", "Planned"),
]

SUMMARY = [
    (1, "Mon Apr 27 to Fri May 1", "Planning & Foundations",
     "Agreed what to build and how.", "",
     "Architecture and tech stack signed off."),
    (2, "Mon May 4 to Fri May 8", "Scanner & Crawler",
     "Reading pages and whole sites into Excel.", "",
     "A user can scan any site and see every field listed in Excel."),
    (3, "Mon May 11 to Fri May 15", "Runner & Self-Healing",
     "Running tests and recovering when pages change.", "",
     "Tests still finish when the form changes."),
    (4, "Mon May 18 to Fri May 22", "Scenario Recording",
     "Record once, replay later.", "Fri May 22",
     "A non technical user can record once and replay reliably later."),
    (5, "Mon May 25 to Fri May 29", "AI Test Cases, Part 1",
     "Generate happy-path variants from a recording.", "Fri May 29",
     "One recording produces at least five working tests without manual edits."),
    (6, "Mon Jun 1 to Fri Jun 5", "AI Test Cases, Part 2",
     "Negative cases, review screen, reports.", "Fri Jun 5",
     "A user goes from one recording to approved and executed AI tests in one sitting."),
]

DEMOS = [
    ("Sprint 4", "Fri May 22",
     "Record a fresh scenario live, save, reload, and replay on a slightly changed page."),
    ("Sprint 5", "Fri May 29",
     "Generate five or more variants from one recording and run them."),
    ("Sprint 6", "Fri Jun 5",
     "Record once, generate cases, review and approve, run them, and show the report."),
]

RISKS = [
    ("Risk", "AI quality on negative cases may be weak.",
     "If output is poor in Sprint 6, ship review and reports only and drop negative cases."),
    ("Risk", "Scenario detail screen is already large.",
     "May need a short cleanup before adding the review tab."),
    ("Risk", "Plain English failure reasons is the most ambitious item.",
     "If it slips, the Jun 5 demo still works without it."),
    ("Assumption", "Single primary developer (Aditya) for the full window.", ""),
    ("Assumption", "Local AI model stays available and performant on the dev machine.", ""),
]

# ---------- styling ----------
THIN = Side(border_style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SPRINT_FILLS = {
    1: "DDEBF7",
    2: "E2EFDA",
    3: "FFF2CC",
    4: "FCE4D6",
    5: "E4DFEC",
    6: "D9E1F2",
}
MILESTONE_FILL = PatternFill("solid", fgColor="C00000")
MILESTONE_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
SUCCESS_FONT = Font(italic=True, size=10, color="375623")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ===== Sheet 1: Plan Overview =====
ws = wb.active
ws.title = "Plan Overview"

ws["A1"] = "Self-Healing Automation: 6 Sprint Delivery Plan"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")

ws["A2"] = "Window: Mon Apr 27, 2026 to Fri Jun 5, 2026.  Owner: Aditya Chatterjee.  Working days only (Mon to Fri)."
ws["A2"].font = SUBTITLE_FONT
ws.merge_cells("A2:F2")

headers = ["Sprint", "Date Range", "Phase", "Focus", "Demo Date", "Success Criteria"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header(ws, 4, len(headers))

for idx, (sp, dr, ph, focus, demo, success) in enumerate(SUMMARY, start=5):
    ws.cell(row=idx, column=1, value=f"Sprint {sp}")
    ws.cell(row=idx, column=2, value=dr)
    ws.cell(row=idx, column=3, value=ph)
    ws.cell(row=idx, column=4, value=focus)
    ws.cell(row=idx, column=5, value=demo)
    ws.cell(row=idx, column=6, value=success)
    fill = PatternFill("solid", fgColor=SPRINT_FILLS[sp])
    for c in range(1, 7):
        cell = ws.cell(row=idx, column=c)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

autosize(ws, [10, 26, 26, 38, 14, 50])
for r in range(5, 5 + len(SUMMARY)):
    ws.row_dimensions[r].height = 38

# ===== Sheet 2: Detailed Tasks =====
ws2 = wb.create_sheet("Detailed Tasks")
ws2["A1"] = "Detailed Task Plan"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:H1")

ws2["A2"] = "Every sub-task by sprint. Demo rows highlighted in red. Success criteria shown under each sprint."
ws2["A2"].font = SUBTITLE_FONT
ws2.merge_cells("A2:H2")

cols = ["Sprint", "Phase", "Date Range", "Sub-Task", "Description", "Demo Date", "Owner", "Status"]
for i, h in enumerate(cols, start=1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(cols))

row_cursor = 5
last_sprint = None
for row in ROWS:
    sp_num = row[0]

    if last_sprint is not None and sp_num != last_sprint:
        success_text = next(s[5] for s in SUMMARY if s[0] == last_sprint)
        ws2.cell(row=row_cursor, column=1, value=f"Sprint {last_sprint} success criteria")
        ws2.cell(row=row_cursor, column=2, value=success_text)
        ws2.merge_cells(start_row=row_cursor, start_column=2,
                        end_row=row_cursor, end_column=len(cols))
        fill = PatternFill("solid", fgColor="EDEDED")
        for c in range(1, len(cols) + 1):
            cell = ws2.cell(row=row_cursor, column=c)
            cell.fill = fill
            cell.border = BORDER
            cell.font = SUCCESS_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.row_dimensions[row_cursor].height = 22
        row_cursor += 1

    for i, v in enumerate(row, start=1):
        cell_value = (f"Sprint {v}" if i == 1 else v)
        cell = ws2.cell(row=row_cursor, column=i, value=cell_value)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    is_milestone = row[3] == "MILESTONE DEMO"
    if is_milestone:
        for i in range(1, len(cols) + 1):
            ws2.cell(row=row_cursor, column=i).fill = MILESTONE_FILL
            ws2.cell(row=row_cursor, column=i).font = MILESTONE_FONT
    else:
        fill = PatternFill("solid", fgColor=SPRINT_FILLS[sp_num])
        for i in range(1, len(cols) + 1):
            ws2.cell(row=row_cursor, column=i).fill = fill

    ws2.row_dimensions[row_cursor].height = 32
    last_sprint = sp_num
    row_cursor += 1

success_text = next(s[5] for s in SUMMARY if s[0] == last_sprint)
ws2.cell(row=row_cursor, column=1, value=f"Sprint {last_sprint} success criteria")
ws2.cell(row=row_cursor, column=2, value=success_text)
ws2.merge_cells(start_row=row_cursor, start_column=2,
                end_row=row_cursor, end_column=len(cols))
fill = PatternFill("solid", fgColor="EDEDED")
for c in range(1, len(cols) + 1):
    cell = ws2.cell(row=row_cursor, column=c)
    cell.fill = fill
    cell.border = BORDER
    cell.font = SUCCESS_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
ws2.row_dimensions[row_cursor].height = 22

autosize(ws2, [10, 24, 24, 20, 55, 14, 12, 14])
ws2.freeze_panes = "A5"

# ===== Sheet 3: Demo Schedule =====
ws3 = wb.create_sheet("Demo Schedule")
ws3["A1"] = "Upcoming Demos"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:C1")
ws3["A2"] = "Every demo lands on a Friday."
ws3["A2"].font = SUBTITLE_FONT
ws3.merge_cells("A2:C2")

dcols = ["Sprint", "Demo Date", "What we showcase"]
for i, h in enumerate(dcols, start=1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(dcols))

for idx, (sp, d, what) in enumerate(DEMOS, start=5):
    ws3.cell(row=idx, column=1, value=sp)
    ws3.cell(row=idx, column=2, value=d)
    ws3.cell(row=idx, column=3, value=what)
    sprint_num = int(sp.split()[-1])
    fill = PatternFill("solid", fgColor=SPRINT_FILLS[sprint_num])
    for c in range(1, 4):
        cell = ws3.cell(row=idx, column=c)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if c == 2:
            cell.font = Font(bold=True)

autosize(ws3, [12, 16, 80])
for r in range(5, 5 + len(DEMOS)):
    ws3.row_dimensions[r].height = 30

# ===== Sheet 4: Risks =====
ws4 = wb.create_sheet("Risks & Assumptions")
ws4["A1"] = "Risks and Assumptions"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:C1")

hdr = ["Type", "Item", "Mitigation"]
for i, h in enumerate(hdr, start=1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, len(hdr))

for idx, r in enumerate(RISKS, start=4):
    for i, v in enumerate(r, start=1):
        cell = ws4.cell(row=idx, column=i, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    type_fill = {
        "Risk": "FCE4D6",
        "Assumption": "E2EFDA",
    }.get(r[0], "FFFFFF")
    fill = PatternFill("solid", fgColor=type_fill)
    for i in range(1, 4):
        ws4.cell(row=idx, column=i).fill = fill

autosize(ws4, [14, 55, 65])
for r in range(4, 4 + len(RISKS)):
    ws4.row_dimensions[r].height = 32

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
print(f"Detailed task rows: {len(ROWS)}")
