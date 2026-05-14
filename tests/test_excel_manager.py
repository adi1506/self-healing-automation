import os
import pytest
from openpyxl import load_workbook
from core.excel_manager import ExcelManager


@pytest.fixture
def tmp_excel(tmp_path):
    """Provide a temporary directory for Excel files."""
    return tmp_path


@pytest.fixture
def manager(tmp_excel):
    """Create an ExcelManager instance with a temp data directory."""
    return ExcelManager(data_dir=str(tmp_excel))


class TestSanitizeURL:
    def test_simple_url(self, manager):
        result = manager.sanitize_url("https://example.com/form")
        assert result == "example_com_form"

    def test_url_with_special_chars(self, manager):
        result = manager.sanitize_url("https://app.example.com/page?id=1&tab=2")
        assert result == "app_example_com_page_id_1_tab_2"


class TestGetExcelPath:
    def test_returns_path_based_on_url(self, manager):
        path = manager.get_excel_path("https://example.com/form")
        assert path.endswith("example_com_form.xlsx")


class TestCreateNewExcel:
    def test_creates_file_with_all_sheets(self, manager):
        elements = [
            {
                "sno": 1,
                "element_name": "First Name",
                "element_type": "input-text",
                "locator_id": "#firstName",
                "locator_name": "firstName",
                "locator_css": "form > div:nth-child(1) > input",
                "locator_xpath": '//*[@id="firstName"]',
                "locator_data_testid": "first-name-input",
                "locator_label": "First Name",
                "placeholder": "Enter first name",
                "available_options": "",
                "current_value": "",
                "status": "NEW",
                "change_details": "",
                "healed_by": "",
            }
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        assert os.path.exists(path)

        wb = load_workbook(path)
        assert "Element Map" in wb.sheetnames
        assert "Test Data" in wb.sheetnames
        assert "Run Results" in wb.sheetnames
        assert "Scan History" in wb.sheetnames
        assert "Heal History" in wb.sheetnames

    def test_element_map_has_correct_data(self, manager):
        elements = [
            {
                "sno": 1,
                "element_name": "First Name",
                "element_type": "input-text",
                "locator_id": "#firstName",
                "locator_name": "firstName",
                "locator_css": "",
                "locator_xpath": '//*[@id="firstName"]',
                "locator_data_testid": "first-name-input",
                "locator_label": "First Name",
                "placeholder": "Enter first name",
                "available_options": "",
                "current_value": "",
                "status": "NEW",
                "change_details": "",
                "healed_by": "",
            }
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Element Map"]
        assert ws.cell(row=2, column=1).value == 1  # S.No
        assert ws.cell(row=2, column=2).value == "First Name"  # Element Name
        assert ws.cell(row=2, column=3).value == "input-text"  # Element Type


class TestTestDataSheet:
    def test_generates_columns_from_editable_elements(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
            {
                "sno": 2, "element_name": "Submit", "element_type": "button",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Test Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "S.No" in headers
        assert "Test Case Name" in headers
        assert "First Name" in headers
        assert "Submit" not in headers


class TestReadElementMap:
    def test_reads_back_saved_elements(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "Email", "element_type": "input-email",
                "locator_id": "#email", "locator_name": "email", "locator_css": "",
                "locator_xpath": "", "locator_data_testid": "email-input",
                "locator_label": "Email", "placeholder": "Enter email",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            }
        ]
        manager.save_element_map("https://example.com/form", elements)
        result = manager.read_element_map("https://example.com/form")
        assert len(result) == 1
        assert result[0]["element_name"] == "Email"
        assert result[0]["locator_id"] == "#email"


class TestTestDataCRUD:
    def test_save_and_read_test_data(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com/form", elements)

        test_rows = [
            {"sno": 1, "test_case_name": "Valid case", "First Name": "John"},
        ]
        manager.save_test_data("https://example.com/form", test_rows)
        result = manager.read_test_data("https://example.com/form")
        assert len(result) == 1
        assert result[0]["First Name"] == "John"


class TestRunResults:
    def test_append_run_result(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com/form", elements)

        run_result = {
            "run_id": "RUN-001",
            "timestamp": "2026-04-01 15:45:00",
            "test_case_name": "Valid case",
            "element_name": "First Name",
            "expected_value": "John",
            "actual_value": "John",
            "status": "PASS",
            "screenshot": "screenshots/run_001.png",
        }
        manager.append_run_result("https://example.com/form", run_result)
        results = manager.read_run_results("https://example.com/form")
        assert len(results) == 1
        assert results[0]["status"] == "PASS"


class TestScanHistory:
    def test_append_scan_history(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com/form", elements)

        scan_entry = {
            "scan_id": "SCAN-001",
            "timestamp": "2026-04-01 15:30:00",
            "total_elements": 12,
            "new": 12,
            "changed": 0,
            "removed": 0,
            "unchanged": 0,
        }
        manager.append_scan_history("https://example.com/form", scan_entry)
        results = manager.read_scan_history("https://example.com/form")
        assert len(results) == 1
        assert results[0]["total_elements"] == 12


class TestHealHistory:
    def test_append_heal_history(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com/form", elements)

        heal_entry = {
            "heal_id": "HEAL-001",
            "timestamp": "2026-04-01 15:30:00",
            "element_name": "First Name",
            "change_type": "CHANGED",
            "change_details": "XPath: /div[2]/input -> /div[3]/input",
            "healed_by": "Level 1 (ID)",
        }
        manager.append_heal_history("https://example.com/form", heal_entry)
        results = manager.read_heal_history("https://example.com/form")
        assert len(results) == 1
        assert results[0]["healed_by"] == "Level 1 (ID)"


class TestMultiRowAppend:
    def test_append_two_run_results(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "X", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com", elements)

        manager.append_run_result("https://example.com", {
            "run_id": "RUN-001", "timestamp": "2026-04-01", "test_case_name": "A",
            "element_name": "X", "expected_value": "1", "actual_value": "1",
            "status": "PASS", "screenshot": "",
        })
        manager.append_run_result("https://example.com", {
            "run_id": "RUN-002", "timestamp": "2026-04-01", "test_case_name": "B",
            "element_name": "X", "expected_value": "2", "actual_value": "2",
            "status": "PASS", "screenshot": "",
        })
        results = manager.read_run_results("https://example.com")
        assert len(results) == 2
        assert results[0]["run_id"] == "RUN-001"
        assert results[1]["run_id"] == "RUN-002"


class TestExcelExists:
    def test_returns_false_for_new_url(self, manager):
        assert manager.excel_exists("https://never-scanned.com") is False

    def test_returns_true_after_save(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "X", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://example.com", elements)
        assert manager.excel_exists("https://example.com") is True


class TestListScannedURLs:
    def test_returns_saved_urls(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "X", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        manager.save_element_map("https://a.com/page1", elements)
        manager.save_element_map("https://b.com/page2", elements)
        urls = manager.list_scanned_urls()
        assert len(urls) == 2


class TestStatusHighlighting:
    def test_changed_elements_highlighted_yellow(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "First Name", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "CHANGED",
                "change_details": "XPath changed", "healed_by": "Level 1",
            },
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Element Map"]
        fill_color = ws.cell(row=2, column=1).fill.start_color.rgb
        assert "FFFF00" in fill_color

    def test_new_elements_highlighted_green(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "New Field", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "NEW",
                "change_details": "", "healed_by": "",
            },
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Element Map"]
        fill_color = ws.cell(row=2, column=1).fill.start_color.rgb
        assert "90EE90" in fill_color

    def test_removed_elements_highlighted_red(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "Old Field", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "REMOVED",
                "change_details": "", "healed_by": "",
            },
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Element Map"]
        fill_color = ws.cell(row=2, column=1).fill.start_color.rgb
        assert "FF6B6B" in fill_color

    def test_unresolved_elements_highlighted_orange(self, manager):
        elements = [
            {
                "sno": 1, "element_name": "Broken Field", "element_type": "input-text",
                "locator_id": "", "locator_name": "", "locator_css": "", "locator_xpath": "",
                "locator_data_testid": "", "locator_label": "", "placeholder": "",
                "available_options": "", "current_value": "", "status": "UNRESOLVED",
                "change_details": "", "healed_by": "",
            },
        ]
        path = manager.save_element_map("https://example.com/form", elements)
        wb = load_workbook(path)
        ws = wb["Element Map"]
        fill_color = ws.cell(row=2, column=1).fill.start_color.rgb
        assert "FFA500" in fill_color


def test_element_map_round_trips_constraint_columns(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "http://example.com/form"
    elements = [
        {
            "sno": 1, "element_name": "Customer Reference", "element_type": "input-text",
            "locator_id": "#custRef", "locator_name": "custRef", "locator_css": "input#custRef",
            "locator_xpath": "//*[@id='custRef']", "locator_data_testid": "",
            "locator_label": "Customer Reference", "placeholder": "",
            "available_options": "", "current_value": "",
            "status": "NEW", "change_details": "", "healed_by": "",
            "pattern": "[A-Z]{4}[0-9]{4}", "title_attr": "4 letters + 4 digits",
            "minlength": "", "maxlength": "8",
            "min": "", "max": "", "autocomplete": "",
            "inputmode": "", "required": True,
            "helper_text": "Format: ABCD1234",
        },
    ]
    em.save_element_map(url, elements)
    read_back = em.read_element_map(url)
    assert read_back[0]["pattern"] == "[A-Z]{4}[0-9]{4}"
    assert read_back[0]["maxlength"] in (8, "8")
    assert read_back[0]["required"] is True
    assert read_back[0]["helper_text"] == "Format: ABCD1234"


def test_test_data_sheet_includes_ai_context_column(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "http://example.com/form"
    elements = [
        {"sno": 1, "element_name": "Email", "element_type": "input-email",
         "locator_id": "", "locator_name": "email", "locator_css": "", "locator_xpath": "",
         "locator_data_testid": "", "locator_label": "Email",
         "placeholder": "", "available_options": "", "current_value": "",
         "status": "NEW", "change_details": "", "healed_by": "",
         "pattern": "", "title_attr": "", "minlength": "", "maxlength": "",
         "min": "", "max": "", "autocomplete": "email", "inputmode": "",
         "required": False, "helper_text": ""},
    ]
    em.save_element_map(url, elements)

    em.save_test_data(url, [
        {"sno": 1, "test_case_name": "Happy path", "ai_context": "Senior citizen",
         "expected_outcome": "success", "Email": "a@b.com"},
    ])
    rows = em.read_test_data(url)
    assert rows[0]["AI Context"] == "Senior citizen"
    assert rows[0]["Test Case Name"] == "Happy path"
    assert rows[0]["Expected Outcome"] == "success"
    assert rows[0]["Email"] == "a@b.com"


def test_expected_outcome_column_backfilled_on_resave(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "http://example.com/form"
    elements = [
        {"sno": 1, "element_name": "Email", "element_type": "input-email",
         "locator_id": "", "locator_name": "email", "locator_css": "", "locator_xpath": "",
         "locator_data_testid": "", "locator_label": "Email",
         "placeholder": "", "available_options": "", "current_value": "",
         "status": "NEW", "change_details": "", "healed_by": "",
         "pattern": "", "title_attr": "", "minlength": "", "maxlength": "",
         "min": "", "max": "", "autocomplete": "email", "inputmode": "",
         "required": False, "helper_text": ""},
    ]
    em.save_element_map(url, elements)

    # Simulate an older sheet by removing the Expected Outcome column.
    from openpyxl import load_workbook
    path = em.get_excel_path(url)
    wb = load_workbook(path)
    ws_td = wb["Test Data"]
    assert ws_td.cell(row=1, column=4).value == "Expected Outcome"
    ws_td.delete_cols(4)
    wb.save(path)

    # Re-saving the element map should backfill the column.
    em.save_element_map(url, elements)
    wb = load_workbook(path)
    headers = [wb["Test Data"].cell(row=1, column=c).value
               for c in range(1, wb["Test Data"].max_column + 1)]
    assert "Expected Outcome" in headers
    assert headers.index("Expected Outcome") == 3  # 0-indexed → column 4


def test_page_context_round_trip(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "http://example.com/form"
    em.save_element_map(url, [])  # ensure workbook exists
    em.save_page_context(url, {
        "title": "Customer Onboarding",
        "h1": "Sign up",
        "first_paragraph": "Welcome to FINN bank.",
    })
    ctx = em.read_page_context(url)
    assert ctx["title"] == "Customer Onboarding"
    assert ctx["h1"] == "Sign up"
    assert ctx["first_paragraph"] == "Welcome to FINN bank."

def test_page_context_returns_empty_when_unset(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "http://example.com/form"
    em.save_element_map(url, [])
    assert em.read_page_context(url) == {"title": "", "h1": "", "first_paragraph": ""}


def test_run_results_round_trip_with_page_index(tmp_path):
    em = ExcelManager(data_dir=str(tmp_path))
    url = "https://e.com/login"
    em.append_run_result(url, {
        "run_id": "abc123", "timestamp": "2026-05-14T10:00:00",
        "test_case_name": "Multi-page login", "row_label": "Page 1",
        "element_name": "email", "expected_value": "a@b.co",
        "actual_value": "a@b.co", "status": "PASS",
        "screenshot": "", "page_index": 0,
    })
    em.append_run_result(url, {
        "run_id": "abc123", "timestamp": "2026-05-14T10:00:05",
        "test_case_name": "Multi-page login", "row_label": "Page 2",
        "element_name": "phone", "expected_value": "1234",
        "actual_value": "1234", "status": "PASS",
        "screenshot": "", "page_index": 1,
    })
    rows = em.read_run_results(url)
    assert len(rows) == 2
    assert rows[0]["page_index"] in (0, "0")
    assert rows[1]["page_index"] in (1, "1")


def test_run_results_back_compat_without_page_index(tmp_path):
    """Existing callers (single-page runs) don't pass page_index. The
    column must default to 0 / empty without crashing."""
    em = ExcelManager(data_dir=str(tmp_path))
    em.append_run_result("https://e.com/x", {
        "run_id": "r", "timestamp": "t", "test_case_name": "tc",
        "row_label": "", "element_name": "n",
        "expected_value": "e", "actual_value": "a", "status": "PASS",
        "screenshot": "",
    })
    rows = em.read_run_results("https://e.com/x")
    assert len(rows) == 1
